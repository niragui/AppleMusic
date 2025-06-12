from typing import Union, List, Tuple

from openpyxl.styles import Font

import datetime

import os
import sys

from .utils import construct_record_id, split_record_id
from .constants import TITLE_FIELD, CREDITS_FIELD, ALBUM_FIELD, ID_FIELD, MATCHING_VALUES_LENGTH
from .exceptions import InvalidID, RecordNotSet, InvalidMatch

sys.path.append(os.getcwd()+"/../..")

from FileData.filerecord import FileRecord

from Utils.artists import ArtistName
from Utils.utils import TAGS_FILE
from Utils.utils import ordinal
from Utils.tags import Tags


GREEN = "70AD47"
RED = "FF0000"
ORANGE = "ED7D31"
BLUE = "4472C4"
WHITE = "FFFFFF"
BLACK = "000000"

POSITIVE_ARROW = "↑"
NEGATIVE_ARROW = "↓"
NO_MOVE_ARROW = "="


RECORD_FIELDS = [TITLE_FIELD, CREDITS_FIELD, ALBUM_FIELD, ID_FIELD]
TITLE_POSITION = 0
CREDITS_POSITION = 1


class DailyChartRecord(FileRecord):
    def __init__(self,
                 date: datetime.date,
                 position: int):
        super().__init__(date, position)
        self.record_name = None
        self.artist_names = None
        self.song_album = None
        self.song_id = None

        self.set = False

    def is_set(self):
        """
        Checks if the information of the record is set
        """
        if self.set:
            return True

        record_set = True
        for field in RECORD_FIELDS:
            if not hasattr(self, field):
                record_set = False
                break

        self.set = record_set
        return record_set

    @property
    def title(self):
        """
        Get the title of the record
        """
        title = getattr(self, TITLE_FIELD, None)

        if title is None:
            raise RecordNotSet("Can't Get Title Of An Unset Record")

        return title

    @property
    def credits(self):
        """
        Get the credits of the record
        """
        credits = getattr(self, CREDITS_FIELD, None)

        if credits is None:
            raise RecordNotSet("Can't Get Credits Of An Unset Record")

        return credits

    #TODO: Check if using ID_FIELD isn't better
    def get_record_id(self):
        """
        Return a string as id to identify the record
        """
        if not self.is_set():
            raise RecordNotSet("Can't Get Record ID Of An Unset Record")

        title = self.title
        credits = self.credits

        record_id = construct_record_id(title, credits)

        return record_id
    def match(self,
              values: Union[List[str], Tuple[str]]):
        """
        Check if the record matches the values sent

        Parameters:
            - values: List or Tuple of [Title, Credits]
        """
        if len(values) > MATCHING_VALUES_LENGTH:
            raise InvalidMatch(f"Match Values Must Be {MATCHING_VALUES_LENGTH} Not More [{len(values)}]")
        elif len(values) < MATCHING_VALUES_LENGTH:
            raise InvalidMatch(f"Match Values Must Be {MATCHING_VALUES_LENGTH} Not Less [{len(values)}]")

        title = self.title
        credits = self.credits

        same_title = title == values[TITLE_POSITION]
        same_credits = credits == values[CREDITS_POSITION]

        return same_title and same_credits

    def get_match_elements(self):
        """
        Return The elements needed to Match The Record
        Should Be Implemented By Each Type Of Record
        """
        matching_values = []

        matching_values.append(self.title)
        matching_values.append(self.credits)

        return matching_values

    def get_writing_values(self):
        """
        Return A List Of The Elements To Write
        In The Database File Of The Record
        Should Be Implemented By Each Type Of Record
        """
        values = []
        for field in RECORD_FIELDS:
            value = getattr(self, field, None)
            if value is None:
                raise RecordNotSet("Can't Get Writing Values Of An Unset Record")

            values.append(value)

        return values


    def is_by_artist(self, artist):
        credits = ArtistName(self.artist_names)

        return credits.is_artist_part(artist)



    def get_taged_name(self):
        artists_names = self.artist_names
        artist = ArtistName(artists_names)
        tags = Tags(TAGS_FILE)
        credits = tags.get_tagged_name(artist)

        return credits

    def get_change_str(self, last_instance, total_days):
        if total_days == 1:
            return "NEW"

        dates_gap = (self.date - last_instance.date).days
        if abs(dates_gap) > 1:
            return "RE"

        change = last_instance.position - self.position

        if change == 0:
            return "="
        elif change < 0:
            return str(change)
        else:
            return "+" + str(change)

    def get_record_text(self, peak, peak_dates, total_days, last_instance):
        change = self.get_change_str(last_instance, total_days)
        text = f"#{self.position} ({change}) "

        text += f"{self.record_name} "
        tagged_name = self.get_taged_name()
        text += tagged_name

        if total_days == 1:
            return text

        text += " "

        if peak == self.position and peak_dates == 1:
            text += "*NEW PEAK*"
            return text

        if peak < self.position:
            text += f"*peak #{peak}"
            if peak_dates > 1:
                text += f" for {peak_dates} days"
            text += "*"
            return text

        ordinal_str = ordinal(peak_dates)
        dates_gap = (self.date - last_instance.date).days

        if abs(dates_gap) > 1 or last_instance.position != self.position:
            text += "*re-peak, "
        else:
            text += "*"
        text += f"{ordinal_str} day at #{peak}*"

        return text

    def get_image_change(self, last_instance, total_days):
        if total_days == 1:
            return "NEW", BLUE

        dates_gap = (self.date - last_instance.date).days
        if abs(dates_gap) > 1:
            return "RE", ORANGE

        change = last_instance.position - self.position

        if change == 0:
            return "=", BLACK
        elif change < 0:
            return NEGATIVE_ARROW + str(abs(change)), RED
        else:
            return POSITIVE_ARROW + str(change), GREEN

    def add_to_report(self, sheet, peak, peak_dates, total_days, last_instance, row):
        cell = sheet.cell(row=row, column=1)  # Position
        cell.value = self.position

        change, color = self.get_image_change(last_instance, total_days)
        cell = sheet.cell(row=row, column=2)  # Change
        cell.value = change
        font = Font(color=color)
        cell.font = font

        cell = sheet.cell(row=row, column=4)  # Song
        cell.value = self.record_name

        cell = sheet.cell(row=row, column=5)  # Credits
        cell.value = self.artist_names

        cell = sheet.cell(row=row, column=6)  # Peak
        cell.value = peak

        cell = sheet.cell(row=row, column=7)  # Peak Days
        peak_ord = ordinal(peak_dates)
        if peak == self.position and peak_dates == 1:
            value = "NEW PEAK"
        elif self.position == peak:
            value = peak_ord
        else:
            value = peak_dates
        cell.value = value

        cell = sheet.cell(row=row, column=8)  # Total Days
        cell.value = total_days

        if last_instance is None:
            return False

        is_re = (self.date - last_instance.date).days > 1
        if is_re:
            cell = sheet.cell(row=row, column=9)  # Total Days
            cell.value = f"First Since: {last_instance.date}"

        return is_re

    def get_image_link(self, albums_image, size):
        link = albums_image.get(self.song_album, None)

        if link is None:
            return None

        link = link.replace("{w}", str(size))
        link = link.replace("{h}", str(size))
        link = link.replace("{f}", "jpg")

        return link

    def is_by_artist(self, artist):
        credits = ArtistName(self.artist_names)

        return credits.is_artist_part(artist)

